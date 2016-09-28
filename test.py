#-*- coding: utf-8 -*-
'''
@author: birdlin
'''
import re
from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import Workbook
from builtins import int

wb = Workbook()
ws = wb.active
row_index = 2
column_index = 1

url = 'http://www.taifex.com.tw/chinese/3/3_1_1_tbl.asp'
list = []
with urlopen(url) as response:
    soup = BeautifulSoup(response.read(), 'lxml')

### Find the title
for title in soup.find_all('th', width=re.compile(r'.*')):
    #if title.string is not None:
        list.append(title.get_text())
    #else :
    #    list.append(title.get_text())

ws.append(list)


for tag in (soup.select("tbody:nth-of-type(1)")):
    for number in tag.find_all('td', class_ = re.compile(r'12')) :
        for string in number.stripped_strings :
            ws.cell(row=row_index, column=column_index).value=string
            column_index += 1
            if column_index > 15 :
                row_index += 1
                column_index = 1

wb.save("BS4sample.xlsx")
print ("Grab Finished!")

'''
for number in soup.find_all('td', class_ = re.compile(r'12')):
    for string in number.stripped_strings :
        ws.cell(row=row_index, column=column_index).value=string
        column_index += 1
        if column_index > 15 :
            row_index += 1
            column_index = 1

wb.save("BS4sample.xlsx")
print ("Grab Finished!")
'''
